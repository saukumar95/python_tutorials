import openpyxl
import re

path = "/home/banti/PycharmProjects/python_tutorials/CC exercise.xlsx"

wb_obj = openpyxl.load_workbook(path)
cc_list = []
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

wb = openpyxl.Workbook()
sheet = wb.active


def readworkbook():
    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        cc_list.append(str(cell_obj.value))
    replacenumber(cc_list)


def replacenumber(ccnumber):
    for i in range(1, 50):
        new_string = re.sub('[0-9]', 'X', ccnumber[i], 12)
        cell_obj = sheet.cell(row=i, column=1)
        cell_obj.value = new_string
        print(new_string)


# def write_new_ccnumber(new_string):


# for i in range(2, m_row + 1):


# print(cell_obj.value)


wb.save("/home/banti/Desktop/CC exercises.xlsx")

# print(new_string);
# print(ccnumber[0])
# print(type(ccnumber))

readworkbook()
